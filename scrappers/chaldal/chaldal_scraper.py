"""
Chaldal.com Web Scraper
Scrapes product data from any page on Chaldal.com and outputs JSON format
"""

import json
import sys
import time
import logging
import csv
from typing import List, Dict, Any, Optional
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ChaldalScraper:
    """Scraper for Chaldal.com product pages"""
    
    BASE_URL = "https://chaldal.com"
    REQUEST_TIMEOUT = 10
    
    def __init__(self, delay_between_requests: float = 1.0):
        """
        Initialize the scraper
        
        Args:
            delay_between_requests: Delay in seconds between requests (default: 1.0)
        """
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        self.delay = delay_between_requests
        self.products = []
    
    def fetch_page(self, url: str) -> Optional[BeautifulSoup]:
        """
        Fetch a page from the URL
        
        Args:
            url: URL to fetch
            
        Returns:
            BeautifulSoup object or None if failed
        """
        try:
            logger.info(f"Fetching: {url}")
            response = self.session.get(url, timeout=self.REQUEST_TIMEOUT)
            response.raise_for_status()
            logger.info(f"Successfully fetched {url}")
            return BeautifulSoup(response.content, 'html.parser')
        except requests.RequestException as e:
            logger.error(f"Error fetching {url}: {e}")
            return None
    
    def extract_products(self, soup: BeautifulSoup) -> List[Dict[str, Any]]:
        """
        Extract product data from the page
        
        Args:
            soup: BeautifulSoup object of the page
            
        Returns:
            List of product dictionaries
        """
        products = []
        
        # Find all product containers - Chaldal uses 'productV2Catalog' class
        product_containers = soup.find_all('div', class_='productV2Catalog')
        
        if not product_containers:
            logger.warning("No product containers found with productV2Catalog class")
            return []
        
        logger.info(f"Found {len(product_containers)} product containers")
        
        for container in product_containers:
            try:
                product = self._extract_product_data(container)
                if product:
                    products.append(product)
            except Exception as e:
                logger.debug(f"Error extracting product: {e}")
                continue
        
        return products
    
    def _extract_product_data(self, container: BeautifulSoup) -> Optional[Dict[str, Any]]:
        """
        Extract individual product data from a container
        
        Args:
            container: BeautifulSoup object of a product container
            
        Returns:
            Product dictionary or None
        """
        try:
            product = {}
            
            # Extract product name from p tag with class 'nameTextWithEllipsis'
            name_elem = container.find('p', class_='nameTextWithEllipsis')
            product['name'] = name_elem.get_text(strip=True) if name_elem else 'N/A'
            
            # Extract prices (handle both discounted and original prices)
            discounted_price_div = container.find('div', class_='productV2discountedPrice')
            
            if discounted_price_div:
                # Get discounted price (first span with amount)
                discounted_span = discounted_price_div.find('span')
                discounted_price = f"৳ {discounted_span.get_text(strip=True)}" if discounted_span else 'N/A'
                
                # Get original price (inside nested price div)
                original_price_div = discounted_price_div.find('div', class_='price')
                if original_price_div:
                    original_span = original_price_div.find('span')
                    original_price = f"৳ {original_span.get_text(strip=True)}" if original_span else 'N/A'
                else:
                    original_price = 'N/A'
                
                product['discounted_price'] = discounted_price
                product['original_price'] = original_price
                product['price'] = discounted_price  # Keep for backward compatibility
            else:
                # Fallback for products without discounted price
                price_span = container.find('div', class_='price')
                if price_span:
                    price_amount = price_span.find('span')
                    price = f"৳ {price_amount.get_text(strip=True)}" if price_amount else 'N/A'
                else:
                    price = 'N/A'
                
                product['price'] = price
                product['discounted_price'] = 'N/A'
                product['original_price'] = 'N/A'
            
            # Extract quantity/size from subText class
            subtext_elem = container.find('div', class_='subText')
            if subtext_elem:
                quantity_elem = subtext_elem.find('span')
                product['quantity'] = quantity_elem.get_text(strip=True) if quantity_elem else 'N/A'
            else:
                product['quantity'] = 'N/A'
            
            # Extract delivery time from deliveryTimeText
            delivery_elem = container.find('div', class_='deliveryTimeText')
            if delivery_elem:
                delivery_span = delivery_elem.find('span')
                product['delivery_time'] = delivery_span.get_text(strip=True) if delivery_span else 'N/A'
            else:
                product['delivery_time'] = 'N/A'
            
            # Extract product image URL
            img_elem = container.find('img')
            if img_elem and img_elem.get('src'):
                product['image_url'] = img_elem.get('src')
            else:
                product['image_url'] = 'N/A'
            
            # For link, we'll construct it from the product name or use a generic link
            # Chaldal doesn't seem to provide direct links in the product container
            product['product_url'] = 'N/A'
            
            return product if product.get('name') != 'N/A' else None
            
        except Exception as e:
            logger.debug(f"Error parsing product data: {e}")
            return None
    
    def scrape_url(self, url: str) -> Dict[str, Any]:
        """
        Scrape a complete URL and return results
        
        Args:
            url: URL to scrape
            
        Returns:
            Dictionary with scraping results
        """
        # Validate URL
        if not url.startswith('http'):
            url = urljoin(self.BASE_URL, url)
        
        if not self._is_valid_url(url):
            logger.error(f"Invalid URL: {url}")
            return {'error': 'Invalid URL', 'url': url}
        
        soup = self.fetch_page(url)
        if not soup:
            return {'error': 'Failed to fetch page', 'url': url}
        
        # Add delay to be respectful to the server
        time.sleep(self.delay)
        
        products = self.extract_products(soup)
        
        result = {
            'url': url,
            'timestamp': datetime.now().isoformat(),
            'total_products': len(products),
            'products': products
        }
        
        logger.info(f"Scraped {len(products)} products from {url}")
        return result
    
    @staticmethod
    def _is_valid_url(url: str) -> bool:
        """
        Validate if URL is valid
        
        Args:
            url: URL to validate
            
        Returns:
            True if valid, False otherwise
        """
        try:
            result = urlparse(url)
            return all([result.scheme, result.netloc])
        except Exception:
            return False


def save_to_json(data: Dict[str, Any], output_file: str = None) -> str:
    """
    Save scraped data to JSON file
    
    Args:
        data: Data to save
        output_file: Output file path (if None, generates one)
        
    Returns:
        Path to saved file
    """
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'chaldal_products_{timestamp}.json'
    
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"Data saved to {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"Error saving to JSON: {e}")
        raise


def save_to_csv(data: Dict[str, Any], output_file: str = None) -> str:
    """
    Save scraped data to CSV file
    
    Args:
        data: Data to save
        output_file: Output file path (if None, generates one)
        
    Returns:
        Path to saved file
    """
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'chaldal_products_{timestamp}.csv'
    
    try:
        products = data.get('products', [])
        
        if not products:
            logger.warning("No products to save")
            return None
        
        # Get all unique keys from all products
        fieldnames = set()
        for product in products:
            fieldnames.update(product.keys())
        fieldnames = sorted(list(fieldnames))
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(products)
        
        logger.info(f"Data saved to {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"Error saving to CSV: {e}")
        raise


def save_to_excel(data: Dict[str, Any], output_file: str = None) -> str:
    """
    Save scraped data to Excel file
    
    Args:
        data: Data to save
        output_file: Output file path (if None, generates one)
        
    Returns:
        Path to saved file
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl is not installed. Install it with: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel export")
    
    if output_file is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'chaldal_products_{timestamp}.xlsx'
    
    try:
        products = data.get('products', [])
        
        if not products:
            logger.warning("No products to save")
            return None
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Get all unique keys from all products
        fieldnames = set()
        for product in products:
            fieldnames.update(product.keys())
        fieldnames = sorted(list(fieldnames))
        
        # Write header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7533CB", end_color="7533CB", fill_type="solid")
        
        for col_idx, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=fieldname)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_idx, product in enumerate(products, 2):
            for col_idx, fieldname in enumerate(fieldnames, 1):
                value = product.get(fieldname, 'N/A')
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, len(fieldnames) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Add metadata sheet
        meta_ws = wb.create_sheet("Metadata")
        meta_ws['A1'] = "URL"
        meta_ws['B1'] = data.get('url', 'N/A')
        meta_ws['A2'] = "Timestamp"
        meta_ws['B2'] = data.get('timestamp', 'N/A')
        meta_ws['A3'] = "Total Products"
        meta_ws['B3'] = data.get('total_products', 0)
        
        meta_ws.column_dimensions['A'].width = 20
        meta_ws.column_dimensions['B'].width = 50
        
        wb.save(output_file)
        logger.info(f"Data saved to {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"Error saving to Excel: {e}")
        raise


def main():
    """Main function to run the scraper"""
    
    if len(sys.argv) < 2:
        print("Usage: python chaldal_scraper.py <URL> [output_file] [--format json|csv|excel]")
        print("\nExamples:")
        print("  python chaldal_scraper.py https://chaldal.com/coffees")
        print("  python chaldal_scraper.py https://chaldal.com/coffees output.json")
        print("  python chaldal_scraper.py https://chaldal.com/coffees output.csv --format csv")
        print("  python chaldal_scraper.py https://chaldal.com/coffees output.xlsx --format excel")
        print("  python chaldal_scraper.py https://chaldal.com/coffees --format csv")
        print("\nSupported formats: json, csv, excel")
        sys.exit(1)
    
    url = sys.argv[1]
    output_file = None
    output_format = 'json'  # default format
    
    # Parse remaining arguments
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '--format':
            if i + 1 < len(sys.argv):
                output_format = sys.argv[i + 1].lower()
                if output_format not in ['json', 'csv', 'excel']:
                    print(f"Error: Invalid format '{output_format}'. Supported: json, csv, excel")
                    sys.exit(1)
                i += 2
            else:
                print("Error: --format requires a value")
                sys.exit(1)
        elif not arg.startswith('--'):
            output_file = arg
            i += 1
        else:
            print(f"Error: Unknown argument '{arg}'")
            sys.exit(1)
    
    scraper = ChaldalScraper(delay_between_requests=1.0)
    result = scraper.scrape_url(url)
    
    if 'error' in result:
        print(json.dumps(result, indent=2))
        sys.exit(1)
    
    # Save in selected format
    if output_format == 'csv':
        saved_file = save_to_csv(result, output_file)
    elif output_format == 'excel':
        saved_file = save_to_excel(result, output_file)
    else:
        saved_file = save_to_json(result, output_file)
    
    # Print summary
    print(f"\n✓ Successfully scraped {result['total_products']} products")
    print(f"✓ Format: {output_format.upper()}")
    print(f"✓ Saved to: {saved_file}")
    print(f"\nData Preview:")
    preview_data = {
        'url': result['url'],
        'timestamp': result['timestamp'],
        'total_products': result['total_products'],
        'first_3_products': result['products'][:3]
    }
    print(json.dumps(preview_data, indent=2, ensure_ascii=False))


if __name__ == '__main__':
    main()
