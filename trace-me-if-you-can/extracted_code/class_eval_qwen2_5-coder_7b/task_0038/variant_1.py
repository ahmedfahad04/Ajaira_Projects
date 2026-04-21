import openpyxl


class ExcelHandler:
    def __init__(self):
        pass

    def load_workbook(self, file_name):
        with openpyxl.load_workbook(file_name) as workbook:
            return workbook.active

    def process_data(self, data, N):
        new_data = []
        for row in data:
            new_row = list(row)
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        return new_data

    def save_workbook(self, data, file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(file_name)
        workbook.close()

    def process_excel_data(self, N, save_file_name):
        data = self.load_workbook(save_file_name).values_only
        if data is None or N >= len(next(data)):
            return 0
        new_data = self.process_data(data, N)
        new_file_name = save_file_name.split('.')[0] + '_processed.xlsx'
        self.save_workbook(new_data, new_file_name)
        return 1, new_file_name
