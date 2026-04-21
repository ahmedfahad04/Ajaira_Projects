import openpyxl


class ExcelHandler:
    def __init__(self):
        pass

    def load_workbook(self, file_name):
        try:
            return openpyxl.load_workbook(file_name).active
        except Exception as e:
            print(f"Failed to load workbook: {e}")
            return None

    def process_data(self, data, N):
        new_data = []
        for row in data:
            new_row = list(row)
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        return new_data

    def save_workbook(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return True
        except Exception as e:
            print(f"Failed to save workbook: {e}")
            return False

    def process_excel_data(self, N, save_file_name):
        data = self.load_workbook(save_file_name).values_only
        if data is None or N >= len(next(data)):
            return 0
        new_data = []
        for row in data:
            new_row = list(row)
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        new_file_name = save_file_name.split('.')[0] + '_processed.xlsx'
        return self.save_workbook(new_data, new_file_name), new_file_name
