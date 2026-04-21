import openpyxl
from contextlib import contextmanager


class ExcelProcessor:
    def __init__(self):
        pass

    @contextmanager
    def _excel_workbook(self, file_name, mode='r'):
        workbook = None
        try:
            if mode == 'r':
                workbook = openpyxl.load_workbook(file_name)
            else:
                workbook = openpyxl.Workbook()
            yield workbook
        except:
            yield None
        finally:
            if workbook:
                workbook.close()

    def read_excel(self, file_name):
        with self._excel_workbook(file_name, 'r') as workbook:
            if workbook is None:
                return None
            sheet = workbook.active
            return [row for row in sheet.iter_rows(values_only=True)]

    def write_excel(self, data, file_name):
        with self._excel_workbook(file_name, 'w') as workbook:
            if workbook is None:
                return 0
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return 0
        
        def append_processed_value(row, index):
            cell_value = row[index]
            processed_value = str(cell_value).upper() if not str(cell_value).isdigit() else cell_value
            return tuple(list(row) + [processed_value])
        
        new_data = [append_processed_value(row, N) for row in data]
        new_file_name = save_file_name.replace('.xlsx', '_process.xlsx').replace('.xls', '_process.xlsx')
        if not new_file_name.endswith('_process.xlsx'):
            new_file_name = save_file_name.split('.')[0] + '_process.xlsx'
        
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name
