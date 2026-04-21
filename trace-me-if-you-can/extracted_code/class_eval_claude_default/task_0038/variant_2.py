import openpyxl


class ExcelProcessor:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def read_excel(self, file_name):
        data = []
        try:
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet = self.workbook.active
            data = list(self.sheet.iter_rows(values_only=True))
            self.workbook.close()
            return data
        except:
            return None

    def write_excel(self, data, file_name):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            [ws.append(row) for row in data]
            wb.save(file_name)
            wb.close()
            return 1
        except:
            return 0

    def _transform_cell_value(self, value):
        return str(value).upper() if not str(value).isdigit() else value

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        
        new_data = []
        for row in data:
            extended_row = list(row) + [self._transform_cell_value(row[N])]
            new_data.append(extended_row)
        
        base_name = save_file_name.rsplit('.', 1)[0]
        new_file_name = f"{base_name}_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name
