import openpyxl
from pathlib import Path


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        try:
            with openpyxl.load_workbook(file_name) as workbook:
                return [row for row in workbook.active.iter_rows(values_only=True)]
        except:
            return None

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        
        processed_data = [
            list(row) + [str(row[N]).upper() if not str(row[N]).isdigit() else row[N]]
            for row in data
        ]
        
        new_file_name = Path(save_file_name).stem + '_process.xlsx'
        success = self.write_excel(processed_data, new_file_name)
        return success, new_file_name
