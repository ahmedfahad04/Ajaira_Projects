import openpyxl


class ExcelProcessor:
    def __init__(self):
        self._data_cache = {}

    def read_excel(self, file_name):
        if file_name in self._data_cache:
            return self._data_cache[file_name]
        
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            rows = []
            row_iter = sheet.iter_rows(values_only=True)
            while True:
                try:
                    row = next(row_iter)
                    rows.append(row)
                except StopIteration:
                    break
            workbook.close()
            self._data_cache[file_name] = rows
            return rows
        except:
            return None

    def write_excel(self, data, file_name):
        workbook = None
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            row_num = 1
            for row in data:
                col_num = 1
                for cell_value in row:
                    sheet.cell(row=row_num, column=col_num, value=cell_value)
                    col_num += 1
                row_num += 1
            workbook.save(file_name)
            return 1
        except:
            return 0
        finally:
            if workbook:
                workbook.close()

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None:
            return 0
        if len(data) == 0 or N >= len(data[0]):
            return 0
        
        new_data = []
        i = 0
        while i < len(data):
            row = data[i]
            new_row = []
            j = 0
            while j < len(row):
                new_row.append(row[j])
                j += 1
            
            target_value = row[N]
            if str(target_value).isdigit():
                new_row.append(target_value)
            else:
                new_row.append(str(target_value).upper())
            
            new_data.append(new_row)
            i += 1
        
        file_parts = save_file_name.split('.')
        new_file_name = file_parts[0] + '_process.xlsx'
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name
